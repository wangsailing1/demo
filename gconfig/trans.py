#! --*-- coding: utf-8 --*--

__author__ = 'sm'


import openpyxl
import hashlib

from gconfig import back_templates
from gconfig import front_templates
from gconfig import analysis
from gconfig import back_contents
from gconfig import front_contents
from gconfig.check import check_reward_counts
from lib.utils.debug import print_log


class TableColError(BufferError):
    """ 列错误

    """
    def __init__(self, message='table col error'):
        self.message = message

    def __str__(self):
        return self.message


def generate_mapping(mapping_config):
    mapping = {}
    for k, v in mapping_config.iteritems():
        if v[0]:
            mapping[v[0]] = k
    return mapping

back_table_mapping_config = generate_mapping(back_contents.mapping_config)
front_table_mapping_config = generate_mapping(front_contents.mapping_config)


def trans_config(filename, sort, xl=None):
    """ 解析配置

    :param filename: 文件名字
    :param sort:
    :param xl:
    :return:
    """
    xl = xl or openpyxl.load_workbook(filename=filename, use_iterators=True)

    result = []

    for sheet in xl.worksheets:
        sheet_title = sheet.title.strip()

        if sort == 'back':
            config_name = back_table_mapping_config.get(sheet_title)
        else:
            config_name = front_table_mapping_config.get(sheet_title)

        if config_name is None:
            continue

        if sort == 'back':
            config_template = getattr(back_templates, sheet_title, None)
        else:
            config_template = getattr(front_templates, sheet_title, None)

        if config_template is None:
            continue

        m = hashlib.md5()
        special_func_name = None
        if isinstance(config_template, tuple):
            config_template, special_func_name = config_template

        sheet_iter = sheet.iter_rows()
        mapping = trans_header(sheet_title, sheet_iter, m, config_template)
        config = trans(sheet_title, mapping, sheet_iter, m, special_func_name)

        m.update(repr(config))
        result.append((config_name, m.hexdigest(), config))

    return result


def trans_header(sheet_title, sheet_iter, md5_obj, config_template):

    headers = {}
    sheet_headers_mapping = {}
    # 中文文字忽略
    sheet_iter.next()

    row = sheet_iter.next()
    for idx, col in enumerate(row):
        sheet_headers_mapping[col.internal_value] = idx

    # for field, (sheet_name, sheet_sort) in config_template.iteritems():
    for field, sheet_list in config_template.iteritems():

        if len(sheet_list) == 2:
            sheet_name, sheet_sort = sheet_list
            sheet_check_func = None
        else:
            sheet_name, sheet_sort, sheet_check_func = sheet_list

        md5_obj.update(repr((field, (sheet_name, sheet_sort))))
        if isinstance(sheet_name, basestring):
            if sheet_name in sheet_headers_mapping:
                headers[field] = (sheet_name, sheet_headers_mapping[sheet_name], sheet_sort, sheet_check_func)
            else:
                raise TableColError('excel "%s" col "%s" not exists, all cols %s' % (sheet_title, sheet_name, sheet_headers_mapping.keys()))
        elif isinstance(sheet_name[0], basestring):
            data = []
            for name in sheet_name:
                if name in sheet_headers_mapping:
                    data.append(sheet_headers_mapping[name])
                else:
                    raise TableColError('excel "%s" col "%s" not exists' % (sheet_title, name))
            headers[field] = (data, sheet_sort, sheet_check_func)
        else:
            data = []
            for name, key_name in sheet_name:
                if name in sheet_headers_mapping:
                    data.append((key_name, sheet_headers_mapping[name]))
                else:
                    raise TableColError('excel "%s" col "%s" not exists' % (sheet_title, name))
            headers[field] = (data, sheet_sort, sheet_check_func)

    return headers


def mapping_func(sheet_sort, value):
    v = analysis.mapping[sheet_sort](value)

    # if value and not v:
    #     raise Exception

    return v

def trans(sheet_title, mapping, sheet_iter, md5_obj, special_func_name):
    if special_func_name:
        special_func = analysis.mapping[special_func_name]
    else:
        special_func = None

    config = {}
    config['check_warning'] = []
    for row_index, row in enumerate(sheet_iter):
        data = {}
        row_data = [r.internal_value for r in row]
        if any(row_data) and not any(row_data[-50:]):     # 防止配置表后边有太多的空列
            raise TableColError(u'excel:%s 后边有太多的空列' % sheet_title)

        if row_data[0] in ['', None]:
            break

        if row_data[0] == '//':
            continue

        md5_obj.update(repr(row_data))
        for field, value in mapping.iteritems():
            msg = ''
            if isinstance(value[0], basestring):
                _, index, sheet_sort, check_func = value
                try:
                    v = row_data[index]
                    if isinstance(sheet_sort, basestring):
                        # v = analysis.mapping[sheet_sort](v)
                        v = mapping_func(sheet_sort, v)
                    else:
                        for s_sort in sheet_sort:
                            # v = analysis.mapping[s_sort](v)
                            v = mapping_func(s_sort, v)
                    # 字段检查函数
                    if check_func is not None:
                        msg = check_func(v)
                        if msg:
                            raise
                    data[field] = v
                    # 奖励数量检查函数
                    if sheet_sort in ['list_3', 'list_4']:
                        msg = check_reward_counts(v)
                        if msg:
                            config['check_warning'].append(('excel:%s row_index:%s field:%s value:%s msg:%s' %
                                                            (sheet_title, row_index, field, row_data[index], msg)))
                except:
                    raise TableColError('1excel:%s row_index:%s field:%s sheet_sort:%s value:%s msg:%s' %
                                        (sheet_title, row_index, field, value, row_data[index], msg))
            else:
                sheets, (sheet_sort, mult_sheet_sort), check_func = value
                args = []
                if not isinstance(sheets[0], tuple):
                    for index in sheets:
                        try:
                            if sheet_sort:
                                # v = analysis.mapping[sheet_sort](row_data[index])
                                v = mapping_func(sheet_sort, row_data[index])
                                args.append(v)
                            else:
                                args.append(row_data[index])
                        except:
                            raise TableColError('excel %s row_index %s field %s sheet_sort %s value %s' %
                                        (sheet_title, row_index, field, value, row_data[index]))
                else:
                    for key_name, index in sheets:
                        try:
                            if sheet_sort:
                                # v = analysis.mapping[sheet_sort](row_data[index])
                                v = mapping_func(sheet_sort, row_data[index])
                                args.append((key_name, v))
                            else:
                                args.append(row_data[index])
                        except:
                            raise TableColError('excel %s row_index %s field %s sheet_sort %s value %s' %
                                        (sheet_title, row_index, field, value, row_data[index]))
                try:
                    # 字段检查函数
                    if check_func is not None:
                        for v in args:
                            msg = check_func(v)
                            if msg:
                                raise
                    data[field] = analysis.mapping[mult_sheet_sort](args)
                    # 奖励数量检查函数
                    for v in args:
                        if sheet_sort in ['list_3', 'list_4']:
                            msg = check_reward_counts(v)
                            if msg:
                                config['check_warning'].append(
                                    ('excel:%s row_index:%s field:%s sheet_sort:%s value:%s msg:%s' %
                                     (sheet_title, row_index, field, value, row_data[index], msg)))
                except:
                    raise TableColError('2excel:%s row_index:%s field:%s sheet_sort:%s value:%s msg:%s' %
                                        (sheet_title, row_index, field, value, args, msg))

        if special_func:
            # if data['uk'] in config:
            #     raise TableColError('excel:%s row_index:%s field:%s sheet_sort:%s value:%s msg:%s' %
            #                         (sheet_title, row_index, 'uk', '', row_data[index], u'id重复'))
            config = special_func(config, data)
        else:
            uk = data.pop('uk')
            if uk in config:
                raise TableColError('3excel:%s row_index:%s field:%s sheet_sort:%s value:%s msg:%s' %
                                    (sheet_title, row_index, 'uk', '', row_data[index], u'id: %s 重复' % uk))
            config[uk] = data

    return config


def import_file(filename):
    """ 导入一个xlsx文件配置

    :param filename:
    :return:
    """
    xl = openpyxl.load_workbook(filename=filename, use_iterators=True)
    trans_config(filename, 'back', xl)
    trans_config(filename, 'front', xl)
